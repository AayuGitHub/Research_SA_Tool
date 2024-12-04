import asyncio  # Import asyncio for managing asynchronous tasks
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from typing import Optional
import csv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import spacy
import re

# Load spaCy model
nlp = spacy.load("en_core_web_sm")  # Replace with your preferred spaCy model

app = FastAPI()


# Helper function to validate text asynchronously
async def validate_text_async(text: str) -> bool:
    """
    Validates the input text using spaCy and regex asynchronously.
    - Ensures no disallowed patterns (e.g., emails, special characters).
    - Ensures meaningful alphabetic content is present.
    """
    email_pattern = r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+"

    # Reject texts with email-like patterns
    if re.search(email_pattern, text):
        raise HTTPException(status_code=400,
                            detail="Text contains disallowed email patterns.")

    # Process text with spaCy asynchronously
    doc = await asyncio.to_thread(nlp, text)

    # Extract alphabetic tokens
    alpha_tokens = [token for token in doc if token.is_alpha]
    if not alpha_tokens:  # No valid alphabetic tokens
        raise HTTPException(status_code=400,
                            detail="Text does not contain any valid words.")

    # Check the proportion of valid alphabetic content
    total_length = len(text)
    clean_length = sum(len(token.text) for token in alpha_tokens)
    if clean_length / total_length < 0.3:  # Less than 30% alphabetic content
        raise HTTPException(
            status_code=400,
            detail=
            "Text contains too many special characters relative to valid content."
        )

    return True


# Asynchronous function to process CSV files
async def process_csv(content: bytes) -> list:
    decoded_content = content.decode("utf-8")
    csv_reader = csv.reader(decoded_content.splitlines())
    return [row for row in csv_reader]


# Asynchronous function to process Excel files
async def process_excel(content: bytes, filename: str) -> list:
    temp_file_path = f"/tmp/{filename}"
    with open(temp_file_path, "wb") as temp_file:
        temp_file.write(content)

    workbook = load_workbook(temp_file_path)
    sheet = workbook.active

    if not isinstance(sheet, Worksheet):
        raise HTTPException(status_code=400,
                            detail="Invalid sheet object in the Excel file.")

    return [list(row) for row in sheet.iter_rows(values_only=True)]


# Endpoint with asynchronous input handling
@app.post("/process-input/")
async def process_input(
    text: Optional[str] = Form(None),  # Optional text query
    file: Optional[UploadFile] = File(None)  # Optional file upload
):
    if not text and not file:
        raise HTTPException(
            status_code=400,
            detail="Please provide either text input, a file, or both.")

    result = {}

    # Process and validate the text asynchronously
    if text:
        try:
            await validate_text_async(text)
            result["text"] = f"Processed text: {text}"
        except HTTPException as e:
            raise e

    # Process the file asynchronously
    if file:
        if not file.filename or file.filename == "":
            raise HTTPException(
                status_code=400,
                detail="Uploaded file is invalid or missing a filename.")

        content = await file.read()

        if file.filename.endswith(".csv"):
            try:
                rows = await process_csv(content)
                result["file"] = {"filename": file.filename, "content": rows}
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Error processing CSV file: {str(e)}")

        elif file.filename.endswith(".txt"):
            try:
                result["file"] = {
                    "filename": file.filename,
                    "content": content.decode("utf-8")
                }
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Error processing TXT file: {str(e)}")

        elif file.filename.endswith(".xlsx"):
            try:
                rows = await process_excel(content, file.filename)
                result["file"] = {"filename": file.filename, "content": rows}
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Error processing Excel file: {str(e)}")

        else:
            raise HTTPException(
                status_code=400,
                detail=
                "Unsupported file type. Only CSV, TXT, or Excel files are allowed."
            )

    return {
        "status": "success",
        "data": result,
        "message": "Input processed successfully"
    }
